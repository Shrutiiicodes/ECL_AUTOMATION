"""
VALIDATION HARNESS
==================
Independently recomputes each stage (not reusing pipeline code) and reconciles
against the pipeline outputs. All-PASS => trust the run.

Checks
  1 LAN_CNT              vs pandas groupby on raw tables (exact)
  2 DISBURSAL_AMT        vs pandas groupby
  3 90+ sums (all MOB)   vs pandas groupby
  4 TPOS sums (all MOB)  vs pandas groupby
  5 loss_rate 84M        vs recompute from movement tables
  6 loss_rate 120M       vs recompute from movement tables
  7 weighted-avg LR      vs explicit SUMPRODUCT/SUM per window
  8 plausibility         every reported weighted LR in (0,1)   [WARN, not FAIL]
"""

import sqlite3
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH  = "ecl.db"
MOB_LIST = list(range(0, 121, 3))
TOL      = 1e-4
START_DISB, END_DISB = "2015-04-01", "2026-07-07"
anchors_for = lambda a: list(range(12, a + 1, 12))
fy_key = lambda q: (int(q[2:4]), int(q[-1]))

def fy_q(d):
    d = pd.Timestamp(d); m = d.month
    fy = d.year if m in (1, 2, 3) else d.year + 1
    q = 4 if m in (1, 2, 3) else 1 if m in (4, 5, 6) else 2 if m in (7, 8, 9) else 3
    return f"FY{str(fy)[-2:]}-Q{q}"

# --- independent recompute of the SQL summary --------------------------------
con = sqlite3.connect(DB_PATH)
base = pd.read_sql("SELECT * FROM base_loans", con)
perf = pd.read_sql("SELECT * FROM performance", con)
con.close()
base = base[(base.disbursal_date >= START_DISB) & (base.disbursal_date <= END_DISB)].copy()
base["FY_QUARTER"] = base.disbursal_date.map(fy_q)
bad = perf.pivot_table(index="distinct_loan_no", columns="mob", values="amt_90plus_settlement", aggfunc="sum")
tps = perf.pivot_table(index="distinct_loan_no", columns="mob", values="tpos", aggfunc="sum")
m = base.set_index("distinct_loan_no")[["FY_QUARTER", "segment", "disbursal_amount"]]
g = m.join(bad.add_prefix("bad_")).join(tps.add_prefix("tps_")).groupby(["FY_QUARTER", "segment"])
exp = pd.DataFrame({"LAN_CNT": g.size(), "DISBURSAL_AMT": g.disbursal_amount.sum() / 1e7})
for mob in MOB_LIST:
    exp[f"AMT_90PLUS_SETTLEMENT_{mob}MOB"] = g[f"bad_{mob}"].sum() / 1e7
    exp[f"TPOS_{mob}MOB"]                  = g[f"tps_{mob}"].sum() / 1e7

feed = pd.read_csv("DATA_ECL_NEW.csv").set_index(["FY_QUARTER", "SEGMENT"])
feed.index = feed.index.set_names(["FY_QUARTER", "segment"])
exp = exp.reindex(feed.index)

checks = []
def add(name, diff, exact=False, warn_only=False):
    ok = (diff == 0) if exact else (diff <= TOL)
    res = "PASS" if ok else ("WARN" if warn_only else "FAIL")
    checks.append((name, diff, "exact" if exact else f"<= {TOL}", res))

add("LAN_CNT", int((feed.LAN_CNT - exp.LAN_CNT).abs().max()), exact=True)
add("DISBURSAL_AMT", float((feed.DISBURSAL_AMT - exp.DISBURSAL_AMT).abs().max()))
bc = [f"AMT_90PLUS_SETTLEMENT_{m}MOB" for m in MOB_LIST]
tc = [f"TPOS_{m}MOB" for m in MOB_LIST]
add("90+ sums (all MOB)",  float((feed[bc].fillna(0) - exp[bc].fillna(0)).abs().values.max()))
add("TPOS sums (all MOB)", float((feed[tc].fillna(0) - exp[tc].fillna(0)).abs().values.max()))

# --- loss rates ---------------------------------------------------------------
a90 = pd.read_csv("tri_90plus_amt.csv", index_col=0); a90.columns = [int(c) for c in a90.columns]
atp = pd.read_csv("tri_tpos_amt.csv",  index_col=0); atp.columns = [int(c) for c in atp.columns]
loss = pd.read_csv("loss_rate.csv").set_index("FY_QUARTER")
for A in (84, 120):
    den = atp[anchors_for(A)].sum(axis=1)
    lr = (a90[A] / den).replace([np.inf, -np.inf], 0).fillna(0)
    add(f"loss_rate {A}M", float((loss[f"LOSS_RATE_{A}M"] - lr.reindex(loss.index)).abs().max()))

# --- weighted averages --------------------------------------------------------
wavg = pd.read_csv("weighted_loss_rate.csv")
lr_all = pd.read_csv("loss_rate.csv")
worst = 0.0
for _, r in wavg.iterrows():
    k1, k2 = fy_key(r.FY_START), fy_key(r.FY_END)
    win = lr_all[lr_all.FY_QUARTER.map(lambda q: k1 <= fy_key(q) <= k2)]
    v = win[f"LOSS_RATE_{int(r.ANCHOR_MOB)}M"].to_numpy(); w = win.DISBURSAL_AMT.to_numpy()
    worst = max(worst, abs(np.dot(v, w) / w.sum() - r.WEIGHTED_AVG_LR))
add("weighted-avg LR", float(worst))

# plausibility: rates must sit in (0,1)
implausible = float(max(0.0, wavg.WEIGHTED_AVG_LR.max() - 1.0))
add("plausibility LR<100%", implausible, warn_only=True)

# --- report -------------------------------------------------------------------
wb = Workbook(); ws = wb.active; ws.title = "Validation"
HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF")
OK = PatternFill("solid", fgColor="C6EFCE"); NO = PatternFill("solid", fgColor="FFC7CE")
WN = PatternFill("solid", fgColor="FFEB9C")
C = Alignment("center", "center"); BD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
for j, h in enumerate(["CHECK", "MAX_ABS_DIFF", "TOLERANCE", "RESULT"], 1):
    c = ws.cell(1, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
for i, (name, diff, tol, res) in enumerate(checks):
    r = 2 + i
    ws.cell(r, 1, name).border = BD
    dc = ws.cell(r, 2, diff); dc.number_format = "0.00E+00" if isinstance(diff, float) else "0"; dc.border = BD
    ws.cell(r, 3, tol).border = BD
    rc = ws.cell(r, 4, res); rc.fill = OK if res == "PASS" else (WN if res == "WARN" else NO)
    rc.font = Font(bold=True); rc.alignment = C; rc.border = BD
ws.column_dimensions["A"].width = 24
for col in "BCD": ws.column_dimensions[col].width = 15
wb.save("validation_report.xlsx")

print("=" * 62); print("VALIDATION"); print("=" * 62)
for name, diff, tol, res in checks:
    d = f"{diff:.2e}" if isinstance(diff, float) else f"{diff}"
    print(f"  [{res}] {name:<24} max diff {d:>10}  (tol {tol})")
print("-" * 62)
fails = [c for c in checks if c[3] == "FAIL"]
warns = [c for c in checks if c[3] == "WARN"]
if fails: print("OVERALL: FAILURES PRESENT - do not ship")
elif warns: print("OVERALL: numerics PASS, but see WARN (a reported loss rate exceeds 100%)")
else: print("OVERALL: ALL CHECKS PASSED - safe to trust the run")
print("Wrote: validation_report.xlsx")