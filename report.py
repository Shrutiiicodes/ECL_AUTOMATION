"""
EXCEL REPORT GENERATION  (live-formula edition, mentor 4+3 tab layout)
=====================================================================
Sheet map (mirrors the manual workbook):

  Summary        headline as-of / ECL (links to Weighted_LR)
  DATA_ECL       raw per-segment feed (values)
  Pivot_ECL      RAW pivot: 90+ amount block + TPOS amount block, each with a
                 Grand Total. Observed cells only; immature cells are BLANK.
                 NO yellow - this is actuals, not projections.
  Chain_Ladder   CHAIN-LADDER triangles with LIVE Excel formulas:
                   90+  as %      (rate = 90+/DISB)
                   TPOS as amount (crores)
                 Mature cells link to Pivot_ECL; immature (yellow) cells carry
                 the bank's =IFERROR(above*SUMPRODUCT(col,DISB)/SUMPRODUCT(col,DISB),0)
                 chain-ladder formula, exactly as in the manual sheet.
  Movements      movement tables (read from Chain_Ladder):
                   TPOS movement amount, TPOS movement %, 90+ movement %
  LossRate       per-quarter loss rate  = 90+@A / SUM(TPOS 12..A)   (from Chain_Ladder)
  Weighted_LR    disbursal-weighted average loss rate per window + final ECL

Every computed cell is a live formula; validation.py reconciles independently.
fullCalcOnLoad makes the workbook show results the moment it opens.

    build_excel(feed, tris, lrr, ecl, path=OUT) -> Workbook
"""

import calendar
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import *      # AS_OF, MOB_LIST, ANCHORS, ANCHOR_MOBS, anchors_for, WINDOWS, HEADLINE, fy_key, OUT

HF   = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF", size=10)
IFL  = PatternFill("solid", fgColor="DDEBF7"); IFONT = Font(bold=True, size=10)
YEL  = PatternFill("solid", fgColor="FFFF00"); WARN = PatternFill("solid", fgColor="FFC7CE")
TOT  = PatternFill("solid", fgColor="C6E0B4")
TITF = Font(bold=True, size=11, color="1F4E78")
CF = Font(size=10); C = Alignment("center", "center"); L = Alignment("left", "center")
BD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
CR, PC, IN = "#,##0.0000", "0.00%", "#,##0"

MOB_COL0 = 3                                        # column of MOB_LIST[0] (A=FY, B=DISB, C=MOB0)
def mob_col(j):        return MOB_COL0 + j
def mob_letter(j):     return get_column_letter(mob_col(j))
def anchor_letter(a):  return mob_letter(MOB_LIST.index(a))

PIVOT = "Pivot_ECL"
WORK  = "Chain_Ladder"


def quarter_end(label):
    fy = 2000 + int(label[2:4]); q = int(label[-1])
    y, m = {1: (fy - 1, 6), 2: (fy - 1, 9), 3: (fy - 1, 12), 4: (fy, 3)}[q]
    return date(y, m, calendar.monthrange(y, m)[1])


def max_mature_mob(label):
    qe = quarter_end(label)
    months = (AS_OF.year - qe.year) * 12 + (AS_OF.month - qe.month)
    v = [m for m in MOB_LIST if m <= months]
    return max(v) if v else -1


is_mature = lambda q, mob: mob <= max_mature_mob(q)


def _hdr_row(ws, r0, prefix):
    for j, h in enumerate(["FY_QUARTER", "DISB_AMT"] + [f"{prefix}{m}MOB" for m in MOB_LIST], 1):
        c = ws.cell(r0, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD


def _widths(ws):
    ws.column_dimensions["A"].width = 12; ws.column_dimensions["B"].width = 12
    for j in range(MOB_COL0, MOB_COL0 + len(MOB_LIST)):
        ws.column_dimensions[get_column_letter(j)].width = 11


# ------------------------------------------------------------------ Pivot_ECL
def write_raw_pivot_block(ws, r0, title, amt, disb):
    """RAW actuals block: observed amounts only (immature -> blank), NO yellow,
    ending in a Grand Total (live =SUM). Returns (first_cohort_row, next_free_row)."""
    tc = ws.cell(r0, 1, title); tc.font = TITF
    _hdr_row(ws, r0 + 1, "")
    top = r0 + 2
    for i, q in enumerate(amt.index):
        r = top + i
        ic = ws.cell(r, 1, q); ic.fill, ic.font, ic.alignment, ic.border = IFL, IFONT, C, BD
        dc = ws.cell(r, 2, round(float(disb.iloc[i]), 6)); dc.number_format, dc.border, dc.alignment = CR, BD, C
        for j, m in enumerate(MOB_LIST):
            val = round(float(amt.iloc[i, j]), 6) if is_mature(q, m) else None   # immature -> blank
            c = ws.cell(r, mob_col(j), val)
            c.number_format, c.font, c.alignment, c.border = CR, CF, C, BD
    gr = top + len(amt.index)
    gc = ws.cell(gr, 1, "Grand Total"); gc.font, gc.fill, gc.border, gc.alignment = Font(bold=True, size=10), TOT, BD, C
    for col_ix in [2] + [mob_col(j) for j in range(len(MOB_LIST))]:
        Lc = get_column_letter(col_ix)
        t = ws.cell(gr, col_ix, f"=SUM({Lc}{top}:{Lc}{gr-1})")
        t.number_format, t.font, t.alignment, t.border, t.fill = CR, Font(bold=True, size=10), C, BD, TOT
    return top, gr + 2


# ------------------------------------------------------------------ Chain_Ladder
def write_chain_ladder_block(ws, r0, title, amt, disb, kind, pivot_top):
    """CHAIN-LADDER triangle with LIVE formulas.
       kind='rate'  -> cells are 90+/DISB (%)   ; mature = pivot/DISB
       kind='amount'-> cells are TPOS (crores)  ; mature = pivot value
    Immature cells (yellow) carry the anchored chain-ladder formula:
       =IFERROR( $X$anchor * SUMPRODUCT(col_top:r-1, DISB) / SUMPRODUCT(col_top:r-2, DISB), 0)
    where $X$anchor is the FIXED deepest-observed cell in the column, so the
    projection does not compound down the column (fix for the deep-MOB blow-up).
    `pivot_top` = first cohort row of the matching block on Pivot_ECL.
    Returns (first_cohort_row, next_free_row)."""
    fmt = PC if kind == "rate" else CR
    tc = ws.cell(r0, 1, title); tc.font = TITF
    _hdr_row(ws, r0 + 1, "")
    top = r0 + 2
    # per-MOB-column anchor: the row of the deepest OBSERVED cohort in that column.
    # Projected cells multiply this FIXED cell (not the cell directly above), so the
    # projection no longer compounds down the column.
    anchor_row = {}
    for j, m in enumerate(MOB_LIST):
        mature_i = [i for i, q in enumerate(amt.index) if is_mature(q, m)]
        anchor_row[j] = (top + mature_i[-1]) if mature_i else None
    for i, q in enumerate(amt.index):
        r = top + i
        pr = pivot_top + i                                  # matching Pivot_ECL row
        ic = ws.cell(r, 1, q); ic.fill, ic.font, ic.alignment, ic.border = IFL, IFONT, C, BD
        dc = ws.cell(r, 2, round(float(disb.iloc[i]), 6)); dc.number_format, dc.border, dc.alignment = CR, BD, C
        for j, m in enumerate(MOB_LIST):
            X = mob_letter(j)
            c = ws.cell(r, mob_col(j))
            if is_mature(q, m):                             # observed -> link to raw pivot
                if kind == "rate":
                    c.value = f"=IFERROR('{PIVOT}'!{X}{pr}/'{PIVOT}'!$B{pr},0)"
                else:
                    c.value = f"='{PIVOT}'!{X}{pr}"
            else:                                           # projected -> chain-ladder formula
                a = anchor_row[j]                            # fixed last-actual row for this column
                num = f"SUMPRODUCT({X}${top}:{X}{r-1},$B${top}:$B{r-1})"
                den = f"SUMPRODUCT({X}${top}:{X}{r-2},$B${top}:$B{r-2})"
                if a is None:                               # no observed cohort in this column
                    c.value = 0
                else:                                       # anchor on the FIXED last actual ($X$a)
                    c.value = f"=IFERROR({X}${a}*{num}/{den},0)"
                c.fill = YEL
            c.number_format, c.font, c.alignment, c.border = fmt, CF, C, BD
    return top, top + len(amt.index) + 1


def build_excel(feed, tris, lrr, ecl, path=OUT):
    a90, atp = tris.a90.copy(), tris.atp.copy()
    disb = tris.disb.copy()
    qtr, wavg = ecl.by_quarter, ecl.wavg
    for df in (a90, atp):
        df.columns = [int(c) for c in df.columns]
    cohorts = list(a90.index)
    n = len(cohorts)

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True

    # ---------------------------------------------------------------- Summary
    ws = wb.active; ws.title = "Summary"
    headline_row = 2 + int(list(wavg.index[wavg.WINDOW == HEADLINE])[0])
    tpos_row, ecl_row = 4 + len(wavg), 4 + len(wavg)
    rows = [
        ("As-of date", str(AS_OF), None),
        ("Cohorts (FY quarters)", n, None),
        ("MOB grid", f"3..120 step 3  ({len(MOB_LIST)} pivot points; 0MOB extracted, not pivoted)", None),
        ("Loss-rate anchors", "84M and 120M", None),
        ("Headline window", HEADLINE, None),
        ("Weighted-avg loss rate", f"='Weighted_LR'!H{headline_row}", PC),
        ("Window total disbursal (cr)", f"='Weighted_LR'!B{tpos_row}", CR),
        ("ECL %", f"='Weighted_LR'!H{headline_row}", PC),
        ("Final-ECL rule", "ECL = disbursal-weighted avg loss rate of the observation window", None),

    ]
    for i, (k, v, fmt) in enumerate(rows):
        r = 1 + i
        kc = ws.cell(r, 1, k); kc.font, kc.fill, kc.border, kc.alignment = IFONT, IFL, BD, L
        vc = ws.cell(r, 2, v); vc.border, vc.alignment = BD, L
        if fmt: vc.number_format = fmt
    ws.column_dimensions["A"].width = 32; ws.column_dimensions["B"].width = 46

    # ---------------------------------------------------------- DATA_ECL
    ws = wb.create_sheet("DATA_ECL")
    for j, h in enumerate(feed.columns, 1):
        c = ws.cell(1, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
    for i, row in feed.iterrows():
        r = 2 + i
        for j, col in enumerate(feed.columns, 1):
            c = ws.cell(r, j, row[col]); c.font, c.border = CF, BD
            if col == "LAN_CNT": c.number_format = IN
            elif col not in ("FY_QUARTER", "SEGMENT"): c.number_format = CR
    ws.freeze_panes = "C2"; ws.column_dimensions["A"].width = 12

    # ------------------------------------------------------------- Pivot_ECL
    ws = wb.create_sheet(PIVOT); _widths(ws)
    p90_top, nxt = write_raw_pivot_block(ws, 1,   "90+ SETTLEMENT (raw actuals, cr)", a90, disb)
    ptp_top, _   = write_raw_pivot_block(ws, nxt, "TPOS (raw actuals, cr)",           atp, disb)
    ws.freeze_panes = ws.cell(3, MOB_COL0)

    # -------------------------------------------------------------- Chain_Ladder
    ws = wb.create_sheet(WORK); _widths(ws)
    w90_top, nxt = write_chain_ladder_block(ws, 1,   "90+% (chain ladder, rate = 90+/DISB)", a90, disb, "rate",   p90_top)
    wtp_top, _   = write_chain_ladder_block(ws, nxt, "TPOS (chain ladder, amount cr)",       atp, disb, "amount", ptp_top)
    ws.freeze_panes = ws.cell(3, MOB_COL0)
    w90_row = {q: w90_top + i for i, q in enumerate(cohorts)}
    wtp_row = {q: wtp_top + i for i, q in enumerate(cohorts)}

    # --------------------------------------------------------- Movements
    ws = wb.create_sheet("Movements")
    def mv_block(r0, title, prefix, kind, src_row, src_is_rate_block):
        """kind: 'amount' (link), 'pct' (link/DISB). src_row maps cohort->Chain_Ladder row.
        src_is_rate_block True means the Chain_Ladder source cells are ALREADY %."""
        fmt = CR if kind == "amount" else PC
        tc = ws.cell(r0, 1, title); tc.font = Font(bold=True, size=10)
        r0 += 1
        for j, h in enumerate(["FY_QUARTER"] + [f"{prefix}{m}MOB" for m in ANCHORS], 1):
            c = ws.cell(r0, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
        for i, q in enumerate(cohorts):
            rr = r0 + 1 + i; sr = src_row[q]
            ic = ws.cell(rr, 1, q); ic.fill, ic.font, ic.border = IFL, IFONT, BD
            for j, a in enumerate(ANCHORS):
                cell = f"'{WORK}'!{anchor_letter(a)}{sr}"
                if kind == "amount":
                    f = f"={cell}"
                elif src_is_rate_block:                     # already % -> link straight
                    f = f"={cell}"
                else:                                       # amount -> % of DISB
                    f = f"=IFERROR({cell}/'{WORK}'!$B{sr},0)"
                c = ws.cell(rr, 2 + j, f)
                c.number_format, c.alignment, c.border = fmt, C, BD
        return r0 + 1 + len(cohorts)
    end = mv_block(1,       "TPOS movement (amount, cr)",     "TPOS_AMT_",   "amount", wtp_row, False)
    end = mv_block(end + 2, "TPOS movement (% of disbursal)", "TPOS_PCT_",   "pct",    wtp_row, False)
    end = mv_block(end + 2, "90+ movement (% of disbursal)",  "90PLUS_PCT_", "pct",    w90_row, True)
    ws.column_dimensions["A"].width = 12

    # -------------------------------------------------------------- LossRate
    ws = wb.create_sheet("LossRate")
    for j, h in enumerate(["FY_QUARTER", "DISB_AMT (weight)", "LOSS_RATE_84M", "LOSS_RATE_120M",
                           "CURRENT_MOB", "CURRENT_TPOS"], 1):
        c = ws.cell(1, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
    lr_row = {}
    for i, rowq in qtr.iterrows():
        q = rowq.FY_QUARTER; rr = 2 + i; lr_row[q] = rr
        w9, wt = w90_row[q], wtp_row[q]
        ic = ws.cell(rr, 1, q); ic.fill, ic.font = IFL, IFONT
        ws.cell(rr, 2, f"='{WORK}'!$B{w9}").number_format = CR       # disbursal weight
        for k, A in enumerate(ANCHOR_MOBS):
            # 90+ amount@A = (Chain_Ladder 90+% @A) * DISB ; TPOS sum = Chain_Ladder TPOS amounts
            num = f"'{WORK}'!{anchor_letter(A)}{w9}*'{WORK}'!$B{w9}"
            den = "+".join(f"'{WORK}'!{anchor_letter(a)}{wt}" for a in anchors_for(A))
            cc = ws.cell(rr, 3 + k, f"=IFERROR(({num})/({den}),0)"); cc.number_format = PC
            if rowq[f"LOSS_RATE_{A}M"] > 1: cc.fill = WARN
        ws.cell(rr, 5, int(rowq.CURRENT_MOB))
        cm = anchor_letter(int(rowq.CURRENT_MOB)) if int(rowq.CURRENT_MOB) in MOB_LIST else mob_letter(0)
        ws.cell(rr, 6, f"='{WORK}'!{cm}{wt}").number_format = CR
        for cc in range(1, 7): ws.cell(rr, cc).alignment = C; ws.cell(rr, cc).border = BD
    ws.column_dimensions["A"].width = 12
    for col in "BCDEF": ws.column_dimensions[col].width = 16

    # ------------------------------------------------------------ Weighted_LR
    ws = wb.create_sheet("Weighted_LR")
    heads = ["WINDOW", "FY_START", "FY_END", "ANCHOR", "N_QTRS", "TOTAL_DISB", "SIMPLE_AVG", "WEIGHTED_AVG"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(1, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
    lr_col = {84: "C", 120: "D"}
    for i, r in wavg.iterrows():
        rr = 2 + i
        k1, k2 = fy_key(r.FY_START), fy_key(r.FY_END)
        win_rows = [lr_row[q] for q in cohorts if k1 <= fy_key(q) <= k2]
        r1, r2 = min(win_rows), max(win_rows)
        LRc = lr_col[int(r.ANCHOR_MOB)]
        disb_rng = f"LossRate!$B{r1}:$B{r2}"
        lr_rng   = f"LossRate!{LRc}{r1}:{LRc}{r2}"
        ws.cell(rr, 1, r.WINDOW); ws.cell(rr, 2, r.FY_START); ws.cell(rr, 3, r.FY_END)
        ws.cell(rr, 4, int(r.ANCHOR_MOB)); ws.cell(rr, 5, r2 - r1 + 1)
        ws.cell(rr, 6, f"=SUM({disb_rng})").number_format = CR
        ws.cell(rr, 7, f"=AVERAGE({lr_rng})").number_format = PC
        wc = ws.cell(rr, 8, f"=IFERROR(SUMPRODUCT({lr_rng},{disb_rng})/SUM({disb_rng}),0)")
        wc.number_format = PC; wc.font = Font(bold=True)
        if r.WEIGHTED_AVG_LR > 1: wc.fill = WARN
        for cc in range(1, 9): ws.cell(rr, cc).alignment = C; ws.cell(rr, cc).border = BD
    r0 = 3 + len(wavg)
    ws.cell(r0, 1, "ECL (%) = weighted-avg LR of headline window").font = Font(bold=True)
    ep = ws.cell(r0, 2, f"=H{headline_row}"); ep.number_format = PC; ep.fill = TOT
    ws.cell(r0 + 1, 1, "Headline window total disbursal (cr)").font = Font(bold=True)
    ws.cell(r0 + 1, 2, f"=F{headline_row}").number_format = CR
    for col, w in zip("ABCDEFGH", [20, 10, 10, 9, 9, 14, 13, 14]):
        ws.column_dimensions[col].width = w

    wb.save(path)
    return wb


if __name__ == "__main__":
    from types import SimpleNamespace
    feed = pd.read_csv("data_ecl.csv")
    a90 = pd.read_csv("tri_90plus_amt.csv", index_col=0);  a90.columns = [int(c) for c in a90.columns]
    atp = pd.read_csv("tri_tpos_amt.csv",  index_col=0);  atp.columns = [int(c) for c in atp.columns]
    disb = feed.groupby("FY_QUARTER").DISBURSAL_AMT.sum().reindex(a90.index)
    qtr  = pd.read_csv("ecl_by_quarter.csv")
    wavg = pd.read_csv("weighted_loss_rate.csv")
    tris = SimpleNamespace(a90=a90, atp=atp, disb=disb)
    ecl  = SimpleNamespace(by_quarter=qtr, wavg=wavg)
    wb = build_excel(feed, tris, SimpleNamespace(), ecl, OUT)
    print("REPORT COMPLETE ->", OUT, "| sheets:", wb.sheetnames)