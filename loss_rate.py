"""
MOVEMENT TABLES + LOSS RATES  (84M and 120M anchors)
====================================================
Consumes the completed amount triangles and produces, per FY_QUARTER:

  movement_tpos    : TPOS at the anchor MOBs
  movement_90plus  : 90+ settlement at the anchor MOBs
  loss table       : loss rate at EACH anchor

Loss rate at anchor A:
    LR_A(q) = 90+(q, A) / SUM( TPOS(q, 12), TPOS(q, 24), ..., TPOS(q, A) )

  84M  -> anchors 12,24,36,48,60,72,84          (7 terms)   == the bank's =I97/SUM(B51:H51)
  120M -> anchors 12,24,...,108,120             (10 terms)

Anchors are MOB LEVELS at those ages (not deltas), matching Final_Workings.
All amounts in crores. DISBURSAL_AMT is carried through because it is the WEIGHT
used for the weighted-average loss rate in the next stage.

-------------------------------------------------------------------------------
This module exposes a PURE function:

    run(tri_90_amt, tri_tpos_amt, feed) -> LossRates(loss, mv90, mvtp)

It reads nothing from disk and writes nothing to disk. The orchestrator passes
DataFrames in and gets DataFrames out. CSV/Excel side effects live only in the
`if __name__ == "__main__"` block below, which exists purely so the phase can
still be run and inspected standalone.
"""

from typing import NamedTuple

import numpy as np
import pandas as pd

from config import *      # ALL_ANCHORS, ANCHOR_MOBS, anchors_for, and (for __main__) TRI_90, TRI_TP, FEED_CSV


class LossRates(NamedTuple):
    loss: pd.DataFrame   # FY_QUARTER, DISBURSAL_AMT, and per-anchor NINETY_PLUS_/TPOS_SUM_12_/LOSS_RATE_
    mv90: pd.DataFrame   # 90+ movement at the anchor MOBs
    mvtp: pd.DataFrame   # TPOS movement at the anchor MOBs


def run(tri_90_amt: pd.DataFrame, tri_tpos_amt: pd.DataFrame, feed: pd.DataFrame) -> LossRates:
    """Compute movement tables and per-anchor loss rates. No I/O."""
    # Triangle columns may arrive as ints (in-memory from chain_ladder) or as
    # strings (round-tripped through CSV in standalone mode). Normalise on copies
    # so we never mutate the caller's frames.
    a90 = tri_90_amt.copy(); a90.columns = [int(c) for c in a90.columns]
    atp = tri_tpos_amt.copy(); atp.columns = [int(c) for c in atp.columns]

    disb = feed.groupby("FY_QUARTER").DISBURSAL_AMT.sum().reindex(a90.index)

    mv90 = a90[ALL_ANCHORS].copy()
    mvtp = atp[ALL_ANCHORS].copy()

    # LOSS RATE PER ANCHOR   (IFERROR -> 0)
    loss = pd.DataFrame({"FY_QUARTER": a90.index, "DISBURSAL_AMT": disb.values})
    for A in ANCHOR_MOBS:
        ancs = anchors_for(A)
        num = a90[A]
        den = atp[ancs].sum(axis=1)
        lr = (num / den).replace([np.inf, -np.inf], 0).fillna(0)
        lr = lr.where(den != 0, 0.0)
        loss[f"NINETY_PLUS_{A}"] = num.values
        loss[f"TPOS_SUM_12_{A}"] = den.values
        loss[f"LOSS_RATE_{A}M"]  = lr.values

    return LossRates(loss=loss, mv90=mv90, mvtp=mvtp)


# Standalone entrypoint: disk I/O + the presentation-only Excel writer.
# None of this runs on import. It exists so `python loss_rate.py` still works.
# The Excel writer is a candidate to move into report.py in the next step.

def _to_excel(res: LossRates, path: str) -> None:
    """Mirror Final_Workings: TPOS movement, 90+ movement, loss rates. Presentation only."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF")
    IFL = PatternFill("solid", fgColor="DDEBF7"); IFONT = Font(bold=True, size=10)
    C = Alignment("center", "center")
    BD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    CR, PC = "#,##0.0000", "0.00%"

    loss, mv90, mvtp = res.loss, res.mv90, res.mvtp
    wb = Workbook(); ws = wb.active; ws.title = "Movements_LossRate"

    def block(ws, r0, prefix, frame):
        for j, h in enumerate(["FY_QUARTER"] + [f"{prefix}{m}MOB" for m in ALL_ANCHORS], 1):
            c = ws.cell(r0, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
        for i, q in enumerate(frame.index):
            rr = r0 + 1 + i
            ic = ws.cell(rr, 1, q); ic.fill, ic.font, ic.border = IFL, IFONT, BD
            for j, m in enumerate(ALL_ANCHORS):
                c = ws.cell(rr, 2 + j, round(float(frame.loc[q, m]), 6))
                c.number_format, c.alignment, c.border = CR, C, BD
        return r0 + 1 + len(frame)

    end = block(ws, 1, "TPOS_", mvtp)
    end = block(ws, end + 2, "90PLUS_", mv90)

    r0 = end + 2
    heads = ["FY_QUARTER", "DISB_AMT (weight)"] + [f"LOSS_RATE_{A}M" for A in ANCHOR_MOBS]
    for j, h in enumerate(heads, 1):
        c = ws.cell(r0, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
    for i, row in loss.iterrows():
        rr = r0 + 1 + i
        ic = ws.cell(rr, 1, row.FY_QUARTER); ic.fill, ic.font, ic.border = IFL, IFONT, BD
        dc = ws.cell(rr, 2, round(row.DISBURSAL_AMT, 4)); dc.number_format = CR
        for k, A in enumerate(ANCHOR_MOBS):
            c = ws.cell(rr, 3 + k, round(row[f"LOSS_RATE_{A}M"], 6)); c.number_format = PC
        for cc in range(2, 3 + len(ANCHOR_MOBS)):
            ws.cell(rr, cc).alignment = C; ws.cell(rr, cc).border = BD
    ws.column_dimensions["A"].width = 12
    for col in "BCD":
        ws.column_dimensions[col].width = 17
    wb.save(path)


def _print_summary(res: LossRates, tri_90_amt: pd.DataFrame) -> None:
    loss, mvtp = res.loss, res.mvtp
    print("=" * 60); print("LOSS RATES COMPLETE"); print("=" * 60)
    print(f"quarters              : {len(loss)}")
    for A in ANCHOR_MOBS:
        col = loss[f"LOSS_RATE_{A}M"]
        print(f"  {A:>3}M anchors {str(anchors_for(A)):<44}")
        print(f"       min/median/max = {col.min():.4%} / {col.median():.4%} / {col.max():.4%}"
              f"   |  >100%: {int((col > 1).sum())}")

    q = "FY18-Q1"
    if q in loss.FY_QUARTER.values:
        r = loss[loss.FY_QUARTER == q].iloc[0]
        hd = float(mvtp.loc[q, anchors_for(84)].sum()); hn = float(tri_90_amt.loc[q, 84])
        print(f"\nhand-check {q} @84M: {hn:.6f} / {hd:.6f} = {hn/hd:.6%}  (engine {r.LOSS_RATE_84M:.6%})")


if __name__ == "__main__":
    a90 = pd.read_csv(TRI_90, index_col=0); a90.columns = [int(c) for c in a90.columns]
    atp = pd.read_csv(TRI_TP, index_col=0); atp.columns = [int(c) for c in atp.columns]
    feed = pd.read_csv(FEED_CSV)

    res = run(a90, atp, feed)

    res.mv90.to_csv("movement_90plus.csv")
    res.mvtp.to_csv("movement_tpos.csv")
    res.loss.to_csv("loss_rate.csv", index=False)
    _to_excel(res, "movements_loss_rate.xlsx")

    _print_summary(res, a90)
    print("\nWrote: movement_90plus.csv, movement_tpos.csv, loss_rate.csv, movements_loss_rate.xlsx")