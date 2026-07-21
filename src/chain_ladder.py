"""
PHASE 2 - TRIANGLE + CHAIN-LADDER ENGINE  (modelling core)
==========================================================
Turns the data_ecl summary into completed run-off triangles for 90+ and
TPOS, filling the immature (yellow) cells with the bank's exact Excel formula.

Triangle layout (mirrors the Chain Ladder sheet)
    rows    = FY_QUARTER cohorts (oldest at top)
    columns = MOB (0,3,...,120)
    cell    = rate = amount(cohort, MOB) / DISBURSAL_AMT(cohort)

Maturity (replaces manual yellow-highlighting)
    A quarter-cohort is MATURE at MOB j only if its whole disbursal window has
    had j months to develop: months_between(quarter_end, AS_OF) >= j.
    Cells beyond that are immature and get projected.

Chain-ladder fill (the bank's exact Excel development-factor formula)
    For an immature cell at (row R, MOB X), where P is the previous MOB column:
        val(R,X) = val(R,P)                                  # same cohort, prev MOB (Excel F46)
                   * SUMPRODUCT(X[top..R-1], DISB[top..R-1]) # this-MOB col, rows above
                   / SUMPRODUCT(P[top..R-1], DISB[top..R-1]) # prev-MOB col, rows above
    i.e. the base is the SAME cohort's previous-MOB value and the multiplier is
    the disbursal-weighted development ratio from the previous MOB to this MOB,
    taken over the mature cohorts ABOVE this row. This reproduces the manual
    sheet's =IFERROR(F46*SUMPRODUCT(G$4:G45,$B$4:$B45)/SUMPRODUCT(F$4:F45,$B$4:$B45),0)
    cell-for-cell (chain_ladder_fill, and the live formulas emitted by report.py).
    The 90+ triangle is filled in RATE space (% cells), then converted to amount
    = rate x disbursal; the TPOS triangle is filled directly in AMOUNT space.
    IFERROR -> 0 when the denominator is 0 or there is no previous column / no
    row above.

This module exposes a PURE function:

    run(feed_raw, segment=SEGMENT) -> Triangles(r90, a90, rtp, atp, mat90, mattp, disb, feed)

`feed_raw` is the data_ecl summary (one row per FY_QUARTER × SEGMENT), as produced
by the SQL phase. `collapse_summary` aggregates it to one row per FY_QUARTER
(optionally filtering to a single segment first). run() reads nothing and writes
nothing. CSV/Excel side effects live only in the `if __name__ == "__main__"` block below.
"""

import calendar
from datetime import date
from typing import NamedTuple

import numpy as np
import pandas as pd

from src.config import *      # AS_OF, MOB_LIST, SEGMENT, and (for __main__) FEED_CSV


class Triangles(NamedTuple):
    r90: pd.DataFrame    # 90+ rate triangle (filled)
    a90: pd.DataFrame    # 90+ amount triangle (filled, crores)
    rtp: pd.DataFrame    # TPOS rate triangle (filled)
    atp: pd.DataFrame    # TPOS amount triangle (filled, crores)
    mat90: pd.DataFrame  # maturity mask for the 90+ build (True = observed)
    mattp: pd.DataFrame  # maturity mask for the TPOS build (identical layout)
    disb: pd.Series      # disbursal per cohort (the chain-ladder weight)
    feed: pd.DataFrame    # the collapsed one-row-per-FY_QUARTER summary


# collapse the per-segment summary to a triangle-ready frame
def collapse_summary(df: pd.DataFrame, segment=None) -> pd.DataFrame:
    """One row per FY_QUARTER (sum amounts + disbursal + count), FY-sorted."""
    df = df.copy()
    if segment is not None:
        df = df[df.SEGMENT == segment].copy()
    amt_cols = [c for c in df.columns if c.startswith(("AMT_90PLUS", "TPOS_"))]
    agg = {"LAN_CNT": "sum", "DISBURSAL_AMT": "sum", **{c: "sum" for c in amt_cols}}
    g = df.groupby("FY_QUARTER").agg(agg)
    key = lambda q: (int(q[2:4]), int(q[-1]))
    return g.reindex(sorted(g.index, key=key))


# MATURITY
def quarter_end(label):                      # 'FY16-Q1' -> 2015-06-30
    fy = 2000 + int(label[2:4]); q = int(label[-1])
    if   q == 1: y, m = fy - 1, 6
    elif q == 2: y, m = fy - 1, 9
    elif q == 3: y, m = fy - 1, 12
    else:        y, m = fy,     3
    return date(y, m, calendar.monthrange(y, m)[1])


def max_mature_mob(label, as_of):
    qe = quarter_end(label)
    months = (as_of.year - qe.year) * 12 + (as_of.month - qe.month)
    valid = [m for m in MOB_LIST if m <= months]
    return max(valid) if valid else -1

def chain_ladder_fill(tri, disb, mature):
    """G46 = IFERROR(F46*SUMPRODUCT(G$4:G45,$B$4:$B45)/SUMPRODUCT(F$4:F45,$B$4:$B45),0)
    Base = same cohort's previous-MOB value (F46); factor = disbursal-weighted
    development ratio from previous MOB (cj-1) to this MOB (cj) over the rows above."""
    R = tri.to_numpy(dtype=float).copy()
    M = mature.to_numpy()
    w = disb.to_numpy(dtype=float)
    n_rows, n_cols = R.shape
    for cj in range(n_cols):                         # left->right: prev MOB column done first
        for ri in range(n_rows):                     # top->bottom: rows above done first
            if M[ri, cj]:
                continue                             # observed cell, keep as-is
            if cj == 0 or ri == 0:
                R[ri, cj] = 0.0                      # no previous MOB / no rows above
                continue
            wa   = w[:ri]                             # $B$top:$B{ri-1}
            num  = np.nansum(R[:ri, cj]     * wa)     # SUMPRODUCT(G$top:G{ri-1}, B)
            den  = np.nansum(R[:ri, cj - 1] * wa)     # SUMPRODUCT(F$top:F{ri-1}, B)
            base = R[ri, cj - 1]                      # F46 = same cohort, previous MOB
            R[ri, cj] = base * num / den if den != 0 else 0.0
    return pd.DataFrame(R, index=tri.index, columns=tri.columns)

# BUILD ONE METRIC'S TRIANGLE
def build(metric_prefix, feed, fill_on="amount"):
    cols = [f"{metric_prefix}{m}MOB" for m in MOB_LIST]
    amt  = feed[cols].copy(); amt.columns = MOB_LIST
    disb = feed["DISBURSAL_AMT"]
    mature = pd.DataFrame(
        [[m <= max_mature_mob(q, AS_OF) for m in MOB_LIST] for q in feed.index],
        index=feed.index, columns=MOB_LIST)
    if fill_on == "rate":
        rate_masked = amt.div(disb, axis=0).where(mature)    # RATE (%) cells, immature -> NaN
        rate_full   = chain_ladder_fill(rate_masked, disb, mature)  # formula runs on % cells (matches Excel)
        amt_full    = rate_full.mul(disb, axis=0)            # amount = rate * disbursal
    else:
        amt_masked = amt.where(mature)                       # immature -> NaN
        amt_full   = chain_ladder_fill(amt_masked, disb, mature)    # formula runs on AMOUNT cells
        rate_full  = amt_full.div(disb, axis=0)              # rate = amount / disbursal
    return rate_full, amt_full, mature, disb


def run(feed_raw: pd.DataFrame, segment=SEGMENT) -> Triangles:
    """Build the completed 90+ and TPOS triangles from the raw SQL summary. No I/O."""
    feed = collapse_summary(feed_raw, segment)
    r90, a90, mat90, disb = build("AMT_90PLUS_SETTLEMENT_", feed, fill_on="rate")
    rtp, atp, mattp, _    = build("TPOS_", feed, fill_on="amount")
    return Triangles(r90=r90, a90=a90, rtp=rtp, atp=atp,
                     mat90=mat90, mattp=mattp, disb=disb, feed=feed)


# Standalone entrypoint: disk I/O + the presentation-only highlighted xlsx.
# None of this runs on import. `python chain_ladder.py` still writes the same
# four CSVs + the yellow-highlighted workbook. The xlsx writer is a candidate
# to move into report.py in a later step.
def _highlighted_xlsx(rate_full, mature, disb, path, title):
    """Visual check: yellow = projected cell. Presentation only."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook(); ws = wb.active; ws.title = title
    HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF")
    YEL = PatternFill("solid", fgColor="FFFF00"); C = Alignment("center", "center")
    BD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    heads = ["FY_QUARTER", "DISB_AMT"] + [f"{m}MOB" for m in MOB_LIST]
    for j, h in enumerate(heads, 1):
        c = ws.cell(1, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
    for i, q in enumerate(rate_full.index):
        r = 2 + i
        ws.cell(r, 1, q).border = BD
        dc = ws.cell(r, 2, round(float(disb.iloc[i]), 4)); dc.number_format = "#,##0.0000"; dc.border = BD
        for j, m in enumerate(MOB_LIST):
            cell = ws.cell(r, 3 + j, round(float(rate_full.iloc[i, j]), 6))
            cell.number_format = "0.00%"; cell.alignment = C; cell.border = BD
            if not mature.iloc[i, j]:
                cell.fill = YEL                              # projected cell
    ws.freeze_panes = "C2"; ws.column_dimensions["A"].width = 11
    wb.save(path)


def _print_summary(tris: Triangles) -> None:
    r90, mat90, disb, feed = tris.r90, tris.mat90, tris.disb, tris.feed
    proj = (~mat90).sum().sum()
    tot  = mat90.size
    print("=" * 60)
    print(f"PHASE 2 COMPLETE   (segment = {SEGMENT or 'ALL'})")
    print("=" * 60)
    print(f"triangle shape        : {r90.shape[0]} cohorts x {r90.shape[1]} MOB")
    print(f"cells                 : {tot}  |  observed {tot-proj}  |  projected {proj}")

    print("\nmax mature MOB per cohort (first 3, last 3):")
    labels = list(feed.index)
    for q in labels[:3] + labels[-3:]:
        print(f"    {q}: {max_mature_mob(q, AS_OF)}")

    # worked example: recompute one projected cell by hand and compare to the engine.
    # The 90+ triangle is filled in RATE space (the formula runs on % cells, then
    # amount = rate x disb), so the hand-check uses r90, the rate triangle.
    r90 = tris.r90
    w   = disb.to_numpy()
    # pick a projected cell whose previous-MOB base is non-zero, so the worked
    # example actually exercises the development factor (not a trivial 0 * factor)
    ri = cj = None
    for j in range(1, len(MOB_LIST)):
        for i in range(1, len(labels)):
            if not mat90.iloc[i, j] and abs(r90.iloc[i, j - 1]) > 1e-9:
                ri, cj = i, j
                break
        if ri is not None:
            break
    base = r90.iloc[ri, cj - 1]                                           # F46: same cohort, previous MOB (rate)
    num  = np.nansum(r90.iloc[:ri, cj].to_numpy()     * w[:ri])           # SUMPRODUCT(this MOB col, DISB) rows above
    den  = np.nansum(r90.iloc[:ri, cj - 1].to_numpy() * w[:ri])           # SUMPRODUCT(prev MOB col, DISB) rows above
    hand = base * num / den if den else 0.0
    print(f"\nworked example (rate)  cell [{labels[ri]}, {MOB_LIST[cj]}MOB]:")
    print(f"    base(prev MOB)={base:.6f}  dev-factor={num/den if den else 0:.6f}")
    print(f"    hand={hand:.8f}   engine={r90.iloc[ri,cj]:.8f}   match={np.isclose(hand, r90.iloc[ri,cj])}")

    # sanity: no NaN left, 90+ rate non-decreasing across MOB (cumulative defaults)
    print(f"\nNaN remaining         : {int(r90.isna().sum().sum())}  (should be 0)")
    nondec = (r90.diff(axis=1).fillna(0) >= -1e-9).all().all()
    print(f"90+ rate non-decreasing across MOB : {nondec}")


if __name__ == "__main__":
    feed_raw = pd.read_csv(FEED_CSV)
    tris = run(feed_raw, SEGMENT)

    tris.r90.to_csv(TRI_90_R); tris.a90.to_csv(TRI_90)
    tris.rtp.to_csv(TRI_TP_R);   tris.atp.to_csv(TRI_TP)
    _highlighted_xlsx(tris.r90, tris.mat90, tris.disb, "triangle_90plus_highlighted.xlsx", "Tri_90plus")

    _print_summary(tris)
    print(f"\nWrote: {TRI_90_R}, {TRI_TP_R}, {TRI_90},")
    print(f"       {TRI_TP}, triangle_90plus_highlighted.xlsx")