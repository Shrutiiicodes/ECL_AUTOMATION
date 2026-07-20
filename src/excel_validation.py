"""
excel_validation.py  -  Tier-1 closed-loop validation of the DELIVERED workbook.

The numeric validation in validation.py reconciles the *pandas* engine against the
DB. It never opens ECL_Report.xlsx, so a correct calculation wired into the wrong
cell (e.g. an ECL% cell pointing at a disbursal reference) passes silently. This
module closes that gap with two levels:

    Level 1  (always runs, no dependency)
        Opens the workbook and asserts STRUCTURE:
          - the sheets we rely on exist
          - the headline ECL cell is a live formula, not a pasted literal
          - the weighted-average cell is the expected SUMPRODUCT/SUM shape
          - the Summary ECL% cell links into Weighted_LR, not somewhere random
        Catches broken references and the 'B6' class of wiring bug.

    Level 2  (runs ONLY if LibreOffice is available)
        Recalculates the workbook headless, reads the computed values back, and
        reconciles them against the Python engine's numbers to TOL.
        Catches formulas that are well-formed but evaluate to the wrong number.
        If LibreOffice is absent it SKIPS (never fails) - so a fresh clone on a
        laptop without LibreOffice still passes Level 1 and reports SKIP for L2.
"""
from __future__ import annotations
import shutil, subprocess, tempfile, os
from openpyxl import load_workbook
from src.config import REPORT_XLSX, HEADLINE, TOL


# ----------------------------------------------------------------- Level 1
def check_structure(xlsx_path: str = REPORT_XLSX) -> list[tuple[str, str, str]]:
    """Return list of (check_name, result, detail). result in PASS/FAIL."""
    out = []
    def add(name, ok, detail=""):
        out.append((name, "PASS" if ok else "FAIL", detail))

    wb = load_workbook(xlsx_path, data_only=False)   # data_only=False -> read FORMULAS

    need = {"Summary", "Weighted_LR"}
    add("required sheets present", need <= set(wb.sheetnames),
        f"missing: {need - set(wb.sheetnames) or 'none'}")

    # headline weighted-average cell must be a live SUMPRODUCT/SUM formula
    wl = wb["Weighted_LR"]
    h2 = str(wl["H2"].value or "")
    add("headline is live formula", h2.startswith("="),
        f"H2={h2[:40]!r}")
    add("headline is SUMPRODUCT/SUM", ("SUMPRODUCT" in h2 and "SUM(" in h2),
        f"H2={h2[:60]!r}")

    # Summary ECL% cell must LINK into Weighted_LR (the 'B6' bug pointed it wrong)
    sm = wb["Summary"]
    ecl_cell = None
    for r in range(1, 20):
        label = str(sm.cell(r, 1).value or "").lower()
        if label.startswith("ecl %") or "ecl (%)" in label:
            ecl_cell = str(sm.cell(r, 2).value or "")
            break
    add("Summary ECL% cell found", ecl_cell is not None)
    add("Summary ECL% links to Weighted_LR",
        ecl_cell is not None and "Weighted_LR" in ecl_cell and ecl_cell.startswith("="),
        f"ref={ecl_cell!r}")
    # the ECL% must reference the weighted-avg column (H), not disbursal (B/F)
    add("Summary ECL% points at weighted-LR column (H)",
        ecl_cell is not None and "H" in ecl_cell.split("!")[-1],
        f"ref={ecl_cell!r}  <- must hit column H, not a disbursal column")
    return out


# ----------------------------------------------------------------- Level 2
def _libreoffice() -> str | None:
    return shutil.which("soffice") or shutil.which("libreoffice")


def check_recalc(xlsx_path: str, python_ecl_pct: float) -> tuple[str, str, str]:
    """(name, result, detail) - result in PASS/FAIL/SKIP. SKIP if no LibreOffice."""
    soffice = _libreoffice()
    if soffice is None:
        return ("recalc vs Python (LibreOffice)", "SKIP",
                "LibreOffice not installed - structural check (L1) still ran")

    tmp = tempfile.mkdtemp()
    try:
        # headless recalc: LibreOffice recomputes all formulas on open+convert
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", tmp, xlsx_path],
            check=True, capture_output=True, timeout=120,
            env={**os.environ, "HOME": tmp},   # isolate profile, avoids stale lock
        )
        recalced = os.path.join(tmp, os.path.basename(xlsx_path))
        wb = load_workbook(recalced, data_only=True)   # data_only=True -> read VALUES
        wl_val = wb["Weighted_LR"]["H2"].value
        if wl_val is None:
            return ("recalc vs Python (LibreOffice)", "FAIL",
                    "recalced H2 is empty - formula did not evaluate")
        diff = abs(float(wl_val) - python_ecl_pct)
        ok = diff <= TOL
        return ("recalc vs Python (LibreOffice)", "PASS" if ok else "FAIL",
                f"workbook={float(wl_val):.6f}  python={python_ecl_pct:.6f}  diff={diff:.2e}")
    except Exception as exc:
        return ("recalc vs Python (LibreOffice)", "FAIL", f"recalc error: {exc}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ----------------------------------------------------------------- driver
def run_excel_validation(python_ecl_pct: float, xlsx_path: str = REPORT_XLSX) -> str:
    """Run both levels, print a report, return overall PASS/FAIL (SKIP != FAIL)."""
    print("=" * 62)
    print("EXCEL-LAYER VALIDATION  (validates the delivered workbook)")
    print("=" * 62)

    rows = check_structure(xlsx_path)                       # Level 1
    rows.append(check_recalc(xlsx_path, python_ecl_pct))    # Level 2

    worst = "PASS"
    for name, res, detail in rows:
        tag = {"PASS": "[PASS]", "FAIL": "[FAIL]", "SKIP": "[SKIP]"}[res]
        print(f"  {tag} {name:<44} {detail}")
        if res == "FAIL":
            worst = "FAIL"
    print("-" * 62)
    print(f"EXCEL VALIDATION: {'ALL PASS' if worst == 'PASS' else 'FAIL - workbook does not match engine'}")
    return worst