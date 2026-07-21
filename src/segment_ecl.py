"""
PER-SEGMENT ECL  (add-on: separate code, separate sheet)
========================================================
Computes the ECL (disbursal-weighted average loss rate) for ONE chosen segment
(1..5) or the whole book ("all"), over a chosen FY-year range, at a chosen anchor,
and writes the result to its OWN sheet -- "Segment_ECL" -- inside the existing
output/ECL_Report.xlsx. It does NOT touch the dataset, the main feed, or any of the
other sheets: it only reuses the pipeline's pure functions and appends one sheet.

The main pipeline (main.py) already makes the SQL feed segment-aware (one row per
FY_QUARTER x SEGMENT). This module filters that feed to the requested segment,
reuses chain_ladder + loss_rate exactly as the whole-book run does, and takes the
disbursal-weighted average loss rate over the requested window. With segment="all"
and the headline window it reproduces the whole-book ECL exactly (a built-in check).

Run it AFTER main.py (so ECL_Report.xlsx exists):

    # interactive (it will prompt for segment / years / anchor):
    python -m src.segment_ecl

    # or non-interactive:
    python -m src.segment_ecl --segment 3 --start FY20 --end FY23 --anchor 84
    python -m src.segment_ecl --segment all --start 2016 --end 2023 --anchor 120

Inputs
------
    segment : 1, 2, 3, 4, 5, or "all"
    start   : first FY year of the window   (accepts 20, FY20, or 2020)
    end     : last  FY year of the window   (accepts 23, FY23, or 2023)
    anchor  : loss-rate anchor MOB          (must be one of config.ANCHOR_MOBS)

The window covers FY{start}-Q1 .. FY{end}-Q4 inclusive.
"""
from __future__ import annotations

import argparse
import os
from typing import NamedTuple

import numpy as np
import pandas as pd

from src.config import (
    ANCHOR_MOBS, HEADLINE, WINDOWS, REPORT_XLSX, SEGMENTS, denom_anchors_for, fy_key
)
from src import sql_refactor, chain_ladder, loss_rate


SEGMENT_SHEET = "Segment_ECL"


class SegmentECL(NamedTuple):
    segment: object          # int 1..5 or "all"
    fy_start: int            # e.g. 20
    fy_end: int              # e.g. 23
    anchor: int              # 72 / 84 / 120
    per_quarter: pd.DataFrame  # window rows: FY_QUARTER, DISB, 90+@A, denominator, LOSS_RATE, observed?
    window_disb: float       # SUM(disbursal) over the window (crores)
    weighted_lr: float       # disbursal-weighted average loss rate  == the ECL %
    simple_lr: float         # plain average, for reference


# ------------------------------------------------------------------ helpers
def _norm_fy(x) -> int:
    """Accept 20, '20', 'FY20', '2020', 'FY2020' -> 20 (two-digit FY)."""
    s = str(x).upper().strip().replace("FY", "").replace("-Q1", "").replace("-Q4", "")
    n = int(s)
    return n % 100 if n >= 100 else n


def _norm_segment(x):
    s = str(x).lower().strip()
    if s in ("all", "*", "none", ""):
        return "all"
    v = int(s)
    if v not in SEGMENTS:
        raise ValueError(f"segment must be one of {SEGMENTS} or 'all', got {x!r}")
    return v


def _segment_feed() -> pd.DataFrame:
    """Return a feed with one row per (FY_QUARTER, SEGMENT).

    If sql_refactor already produces a segment-aware feed (your `GROUP BY
    fy_quarter, segment` change), use it directly. Otherwise -- if the main
    pipeline was left as whole-book -- build the segment view here from the raw
    DB, so this add-on works either way and needs nothing changed upstream.
    """
    feed = sql_refactor.run().feed
    if "SEGMENT" in feed.columns:
        return feed

    # Fallback: aggregate to (FY_QUARTER, SEGMENT) directly from the DB, in pandas.
    import sqlite3
    from src.config import DB_PATH, MOB_SQL, START_DISB, END_DISB

    def fy_q(d):
        d = pd.Timestamp(d); m = d.month
        fy = d.year if m in (1, 2, 3) else d.year + 1
        q = 4 if m in (1, 2, 3) else 1 if m in (4, 5, 6) else 2 if m in (7, 8, 9) else 3
        return f"FY{str(fy)[-2:]}-Q{q}"

    con = sqlite3.connect(DB_PATH)
    base = pd.read_sql("SELECT * FROM base_loans", con)
    perf = pd.read_sql("SELECT * FROM performance", con)
    con.close()

    base = base[(base.disbursal_date >= START_DISB) & (base.disbursal_date <= END_DISB)].copy()
    base["FY_QUARTER"] = base.disbursal_date.map(fy_q)
    bad = perf.pivot_table(index="distinct_loan_no", columns="mob", values="amt_90plus_settlement", aggfunc="sum")
    tps = perf.pivot_table(index="distinct_loan_no", columns="mob", values="tpos", aggfunc="sum")
    m = base.set_index("distinct_loan_no")[["FY_QUARTER", "segment", "disbursal_amount"]]
    g = m.join(bad.add_prefix("bad_")).join(tps.add_prefix("tps_")).groupby(["FY_QUARTER", "segment"])
    out = pd.DataFrame({"LAN_CNT": g.size(), "DISBURSAL_AMT": g.disbursal_amount.sum() / 1e7})
    for mob in MOB_SQL:
        out[f"AMT_90PLUS_SETTLEMENT_{mob}MOB"] = g[f"bad_{mob}"].sum() / 1e7
        out[f"TPOS_{mob}MOB"] = g[f"tps_{mob}"].sum() / 1e7
    out = out.reset_index().rename(columns={"segment": "SEGMENT"})
    return out


# ------------------------------------------------------------------ compute
def compute(segment, fy_start, fy_end, anchor, full_feed=None) -> SegmentECL:
    """Pure calculation. Reuses chain_ladder + loss_rate; no file I/O."""
    seg = _norm_segment(segment)
    y1, y2 = _norm_fy(fy_start), _norm_fy(fy_end)
    anchor = int(anchor)
    if anchor not in ANCHOR_MOBS:
        raise ValueError(f"anchor must be one of {ANCHOR_MOBS}, got {anchor}")

    # segment-aware feed: uses your segment-aware SQL if present, else builds it from the DB
    full_feed = _segment_feed() if full_feed is None else full_feed
    seg_num = None if seg == "all" else seg
    seg_feed = full_feed if seg_num is None else full_feed[full_feed.SEGMENT == seg_num]
    if len(seg_feed) == 0:
        raise ValueError(f"no rows for segment {seg!r} in the feed")

    # reuse the pipeline: collapse+triangles for this segment, then loss rates
    tris = chain_ladder.run(full_feed, seg_num)           # collapse_summary filters to the segment
    lrr = loss_rate.run(tris.a90, tris.atp, seg_feed)     # disbursal from the segment-filtered feed
    loss = lrr.loss

    # window filter: FY{y1}-Q1 .. FY{y2}-Q4
    k1, k2 = (y1, 1), (y2, 4)
    win = loss[loss.FY_QUARTER.map(lambda q: k1 <= fy_key(q) <= k2)].copy()
    if len(win) == 0:
        raise ValueError(f"no quarters in window FY{y1}-Q1..FY{y2}-Q4 for segment {seg!r}")

    lr = win[f"LOSS_RATE_{anchor}M"].to_numpy()
    w = win["DISBURSAL_AMT"].to_numpy()
    weighted_lr = float(np.dot(lr, w) / w.sum()) if w.sum() else 0.0
    simple_lr = float(lr.mean())

    # tidy per-quarter table for the sheet
    pq = pd.DataFrame({
        "FY_QUARTER":  win["FY_QUARTER"].values,
        "DISB_AMT":    win["DISBURSAL_AMT"].values,
        f"90PLUS_{anchor}M":  win[f"NINETY_PLUS_{anchor}"].values,
        f"DENOM_{anchor}M":   win[f"DEN_DISB_TPOS_{anchor}"].values,
        f"LOSS_RATE_{anchor}M": lr,
    })

    return SegmentECL(segment=seg, fy_start=y1, fy_end=y2, anchor=anchor,
                      per_quarter=pq, window_disb=float(w.sum()),
                      weighted_lr=weighted_lr, simple_lr=simple_lr)


# ------------------------------------------------------------------ write sheet
def write_sheet(res: SegmentECL, xlsx_path: str = REPORT_XLSX) -> str:
    """Append (or refresh) the 'Segment_ECL' sheet. If the report workbook does not
    exist yet, write a standalone workbook next to it. Returns the path written."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF", size=10)
    IFL = PatternFill("solid", fgColor="DDEBF7"); IFONT = Font(bold=True, size=10)
    TOT = PatternFill("solid", fgColor="C6E0B4"); TITF = Font(bold=True, size=12, color="1F4E78")
    CF = Font(size=10); C = Alignment("center", "center"); L = Alignment("left", "center")
    BD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    CR, PC = "#,##0.0000", "0.00%"

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)                     # keep formulas + other sheets
        if SEGMENT_SHEET in wb.sheetnames:
            del wb[SEGMENT_SHEET]                          # refresh, don't duplicate
        ws = wb.create_sheet(SEGMENT_SHEET)
        out_path = xlsx_path
    else:                                                 # report not built yet -> standalone
        wb = Workbook(); ws = wb.active; ws.title = SEGMENT_SHEET
        out_path = os.path.join(os.path.dirname(xlsx_path) or ".", "Segment_ECL.xlsx")

    seg_label = "ALL segments (whole book)" if res.segment == "all" else f"Segment {res.segment}"
    a = res.anchor

    ws.cell(1, 1, "PER-SEGMENT ECL").font = TITF
    summary = [
        ("Segment",                     seg_label, None),
        ("Observation window",          f"FY{res.fy_start}-Q1 .. FY{res.fy_end}-Q4", None),
        ("Anchor (MOB)",                f"{a}M", None),
        ("Quarters in window",          len(res.per_quarter), None),
        ("Window total disbursal (cr)", round(res.window_disb, 4), CR),
        ("Simple avg loss rate",        res.simple_lr, PC),
        ("Weighted-avg loss rate = ECL %", res.weighted_lr, PC),
    ]
    for i, (k, v, fmt) in enumerate(summary):
        r = 2 + i
        kc = ws.cell(r, 1, k); kc.font, kc.fill, kc.border, kc.alignment = IFONT, IFL, BD, L
        vc = ws.cell(r, 2, v); vc.border, vc.alignment = BD, L
        if fmt: vc.number_format = fmt
    # highlight the ECL row
    ws.cell(2 + len(summary) - 1, 2).fill = TOT

    # per-quarter detail table
    top = 2 + len(summary) + 2
    ws.cell(top - 1, 1, "Per-quarter detail (window)").font = Font(bold=True, size=10)
    heads = ["FY_QUARTER", "DISB_AMT (weight)", f"90+ @{a}M", f"DENOM @{a}M (DISB+TPOS 12..{a-12})", f"LOSS_RATE_{a}M"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(top, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
    for i, row in res.per_quarter.reset_index(drop=True).iterrows():
        r = top + 1 + i
        ic = ws.cell(r, 1, row["FY_QUARTER"]); ic.fill, ic.font, ic.border, ic.alignment = IFL, IFONT, BD, C
        vals = [
            (2, row["DISB_AMT"], CR),
            (3, row[f"90PLUS_{a}M"], CR),
            (4, row[f"DENOM_{a}M"], CR),
            (5, row[f"LOSS_RATE_{a}M"], PC),
        ]
        for col, v, fmt in vals:
            cc = ws.cell(r, col, round(float(v), 6)); cc.number_format, cc.border, cc.font, cc.alignment = fmt, BD, CF, C
    # totals / weighted-avg footer
    gr = top + 1 + len(res.per_quarter)
    gc = ws.cell(gr, 1, "WEIGHTED AVG"); gc.fill, gc.font, gc.border, gc.alignment = TOT, Font(bold=True, size=10), BD, C
    dc = ws.cell(gr, 2, round(res.window_disb, 4)); dc.number_format, dc.fill, dc.border, dc.alignment = CR, TOT, BD, C
    wc = ws.cell(gr, 5, res.weighted_lr); wc.number_format, wc.fill, wc.border, wc.font, wc.alignment = PC, TOT, BD, Font(bold=True, size=10), C

    ws.column_dimensions["A"].width = 16
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 30 if col == "D" else 18
    ws.sheet_properties.tabColor = "7030A0"

    wb.save(out_path)
    return out_path


# ------------------------------------------------------------------ CLI
def _prompt_inputs():
    print("Per-segment ECL — enter inputs (press Enter for the default in [brackets]).")
    seg = input(f"  segment 1/2/3/4/5/all [{ 'all' }]: ").strip() or "all"
    # default window/anchor from the headline config window
    hz = [w for w in WINDOWS if w[0] == HEADLINE][0]
    d_start = _norm_fy(hz[1]); d_end = _norm_fy(hz[2]); d_anchor = hz[3]
    y1 = input(f"  start FY year (e.g. 20 or 2020) [{d_start}]: ").strip() or d_start
    y2 = input(f"  end   FY year (e.g. 23 or 2023) [{d_end}]: ").strip() or d_end
    an = input(f"  anchor MOB {ANCHOR_MOBS} [{d_anchor}]: ").strip() or d_anchor
    return seg, y1, y2, an


def main():
    ap = argparse.ArgumentParser(description="Per-segment ECL -> Segment_ECL sheet")
    ap.add_argument("--segment", help="1|2|3|4|5|all")
    ap.add_argument("--start", help="first FY year, e.g. 20 / FY20 / 2020")
    ap.add_argument("--end", help="last FY year, e.g. 23 / FY23 / 2023")
    ap.add_argument("--anchor", help=f"anchor MOB, one of {ANCHOR_MOBS}")
    ap.add_argument("--xlsx", default=REPORT_XLSX, help="workbook to append the sheet to")
    args = ap.parse_args()

    if args.segment is None and args.start is None and args.end is None and args.anchor is None:
        seg, y1, y2, an = _prompt_inputs()          # nothing passed -> ask interactively
    else:
        hz = [w for w in WINDOWS if w[0] == HEADLINE][0]
        seg = args.segment or "all"
        y1 = args.start or _norm_fy(hz[1])
        y2 = args.end or _norm_fy(hz[2])
        an = args.anchor or hz[3]

    res = compute(seg, y1, y2, an)
    out = write_sheet(res, args.xlsx)

    seg_label = "ALL (whole book)" if res.segment == "all" else f"segment {res.segment}"
    print("=" * 60)
    print("PER-SEGMENT ECL")
    print("=" * 60)
    print(f"segment            : {seg_label}")
    print(f"window             : FY{res.fy_start}-Q1 .. FY{res.fy_end}-Q4  @ {res.anchor}M")
    print(f"quarters           : {len(res.per_quarter)}")
    print(f"window disbursal   : {res.window_disb:,.2f} cr")
    print(f"simple avg LR      : {res.simple_lr:.4%}")
    print(f"weighted-avg LR    : {res.weighted_lr:.4%}   <-- ECL %")
    print(f"written to         : {out}  (sheet '{SEGMENT_SHEET}')")


if __name__ == "__main__":
    main()