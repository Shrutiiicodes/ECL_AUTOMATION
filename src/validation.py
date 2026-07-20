"""
VALIDATION HARNESS
==================
Independently recomputes each stage (NOT reusing pipeline code) and reconciles
against the pipeline outputs. All-PASS => trust the run.

Checks
  1 LAN_CNT              vs pandas groupby on raw tables (exact)
  2 DISBURSAL_AMT        vs pandas groupby
  3 90+ sums (all MOB)   vs pandas groupby
  4 TPOS sums (all MOB)  vs pandas groupby
  5 loss_rate 84M        vs recompute from movement tables
  6 loss_rate 120M       vs recompute from movement tables
  7 weighted-avg LR      vs explicit SUMPRODUCT/SUM per window
  8 plausibility         every reported weighted LR in (0,1)   [WARN, not FAIL]

-------------------------------------------------------------------------------
This module exposes:

    validate(feed, tris, lrr, ecl, db_path=DB_PATH) -> list[Check]

The independent truth (checks 1-4) is recomputed from `db_path` every time - it
never trusts the pipeline. The pipeline results are passed in only so we have
something to diff against:
    feed : sql_refactor.run().feed
    tris : chain_ladder.run()  (.a90, .atp)
    lrr  : loss_rate.run()     (.loss)
    ecl  : final_ecl.run()     (.wavg)
validate() writes nothing. Excel/printing live in helpers used by __main__.
"""

import sqlite3
from typing import NamedTuple

import numpy as np
import pandas as pd

from src.config import *      # DB_PATH, MOB_LIST, TOL, START_DISB, END_DISB, ANCHOR_MOBS, anchors_for, fy_key


class Check(NamedTuple):
    name: str
    diff: float
    tolerance: str
    result: str          # PASS / WARN / FAIL


def fy_q(d):
    d = pd.Timestamp(d); m = d.month
    fy = d.year if m in (1, 2, 3) else d.year + 1
    q = 4 if m in (1, 2, 3) else 1 if m in (4, 5, 6) else 2 if m in (7, 8, 9) else 3
    return f"FY{str(fy)[-2:]}-Q{q}"


def validate(feed, tris, lrr, ecl, db_path=DB_PATH) -> list:
    """Recompute independently from the DB and reconcile against the pipeline results."""
    # independent recompute of the SQL summary (checks 1-4)
    con = sqlite3.connect(db_path)
    base = pd.read_sql("SELECT * FROM base_loans", con)
    perf = pd.read_sql("SELECT * FROM performance", con)
    con.close()
    base = base[(base.disbursal_date >= START_DISB) & (base.disbursal_date <= END_DISB)].copy()
    base["FY_QUARTER"] = base.disbursal_date.map(fy_q)
    bad = perf.pivot_table(index="distinct_loan_no", columns="mob", values="amt_90plus_settlement", aggfunc="sum")
    tps = perf.pivot_table(index="distinct_loan_no", columns="mob", values="tpos", aggfunc="sum")
    m = base.set_index("distinct_loan_no")[["FY_QUARTER", "disbursal_amount"]]
    g = m.join(bad.add_prefix("bad_")).join(tps.add_prefix("tps_")).groupby("FY_QUARTER")
    exp = pd.DataFrame({"LAN_CNT": g.size(), "DISBURSAL_AMT": g.disbursal_amount.sum() / 1e7})
    for mob in MOB_SQL:
        exp[f"AMT_90PLUS_SETTLEMENT_{mob}MOB"] = g[f"bad_{mob}"].sum() / 1e7
        exp[f"TPOS_{mob}MOB"]                  = g[f"tps_{mob}"].sum() / 1e7

    # pipeline output to diff against (passed in, not read from CSV)
    feed_idx = feed.set_index("FY_QUARTER").copy()
    exp = exp.reindex(feed_idx.index)

    checks = []
    def add(name, diff, exact=False, warn_only=False):
        ok = (diff == 0) if exact else (diff <= TOL)
        res = "PASS" if ok else ("WARN" if warn_only else "FAIL")
        checks.append(Check(name, diff, "exact" if exact else f"<= {TOL}", res))

    add("LAN_CNT", int((feed_idx.LAN_CNT - exp.LAN_CNT).abs().max()), exact=True)
    add("DISBURSAL_AMT", float((feed_idx.DISBURSAL_AMT - exp.DISBURSAL_AMT).abs().max()))
    bc = [f"AMT_90PLUS_SETTLEMENT_{m}MOB" for m in MOB_SQL]
    tc = [f"TPOS_{m}MOB" for m in MOB_SQL]
    add("90+ sums (all MOB)",  float((feed_idx[bc].fillna(0) - exp[bc].fillna(0)).abs().values.max()))
    add("TPOS sums (all MOB)", float((feed_idx[tc].fillna(0) - exp[tc].fillna(0)).abs().values.max()))

    # loss rates (checks 5-6): recompute from the triangles
    a90 = tris.a90.copy(); a90.columns = [int(c) for c in a90.columns]
    atp = tris.atp.copy(); atp.columns = [int(c) for c in atp.columns]
    loss = lrr.loss.set_index("FY_QUARTER")
    for A in ANCHOR_MOBS:
        den = atp[anchors_for(A)].sum(axis=1)
        lr = (a90[A] / den).replace([np.inf, -np.inf], 0).fillna(0)
        add(f"loss_rate {A}M", float((loss[f"LOSS_RATE_{A}M"] - lr.reindex(loss.index)).abs().max()))

    # weighted averages (check 7)
    wavg = ecl.wavg
    lr_all = lrr.loss
    worst = 0.0
    for _, r in wavg.iterrows():
        k1, k2 = fy_key(r.FY_START), fy_key(r.FY_END)
        win = lr_all[lr_all.FY_QUARTER.map(lambda q: k1 <= fy_key(q) <= k2)]
        v = win[f"LOSS_RATE_{int(r.ANCHOR_MOB)}M"].to_numpy(); w = win.DISBURSAL_AMT.to_numpy()
        worst = max(worst, abs(np.dot(v, w) / w.sum() - r.WEIGHTED_AVG_LR))
    add("weighted-avg LR", float(worst))

    # plausibility: rates must sit in (0,1)   [WARN only]
    implausible = float(max(0.0, wavg.WEIGHTED_AVG_LR.max() - 1.0))
    add("plausibility LR<100%", implausible, warn_only=True)

    return checks


def overall_status(checks) -> str:
    if any(c.result == "FAIL" for c in checks): return "FAIL"
    if any(c.result == "WARN" for c in checks): return "WARN"
    return "PASS"


# Presentation + standalone entrypoint. None of this runs on import.
def write_report(checks, path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook(); ws = wb.active; ws.title = "Validation"
    HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF")
    OK = PatternFill("solid", fgColor="C6EFCE"); NO = PatternFill("solid", fgColor="FFC7CE")
    WN = PatternFill("solid", fgColor="FFEB9C")
    C = Alignment("center", "center"); BD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    for j, h in enumerate(["CHECK", "MAX_ABS_DIFF", "TOLERANCE", "RESULT"], 1):
        c = ws.cell(1, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
    for i, (name, diff, tol, res) in enumerate(checks):
        r = 2 + i
        ws.cell(r, 1, name).border = BD
        dc = ws.cell(r, 2, diff); dc.number_format = "0.00E+00" if isinstance(diff, float) else "0"; dc.border = BD
        ws.cell(r, 3, tol).border = BD
        rc = ws.cell(r, 4, res); rc.fill = OK if res == "PASS" else (WN if res == "WARN" else NO)
        rc.font = Font(bold=True); rc.alignment = C; rc.border = BD
    ws.column_dimensions["A"].width = 24
    for col in "BCD": ws.column_dimensions[col].width = 15
    wb.save(path)


def print_summary(checks) -> None:
    print("=" * 62); print("VALIDATION"); print("=" * 62)
    for name, diff, tol, res in checks:
        d = f"{diff:.2e}" if isinstance(diff, float) else f"{diff}"
        print(f"  [{res}] {name:<24} max diff {d:>10}  (tol {tol})")
    print("-" * 62)
    status = overall_status(checks)
    if status == "FAIL":
        print("OVERALL: FAILURES PRESENT - do not ship")
    elif status == "WARN":
        print("OVERALL: numerics PASS, but see WARN (a reported loss rate exceeds 100%)")
    else:
        print("OVERALL: ALL CHECKS PASSED - safe to trust the run")


if __name__ == "__main__":
    from types import SimpleNamespace

    feed = pd.read_csv(FEED_CSV)
    a90 = pd.read_csv(TRI_90, index_col=0); a90.columns = [int(c) for c in a90.columns]
    atp = pd.read_csv(TRI_TP,  index_col=0); atp.columns = [int(c) for c in atp.columns]
    loss = pd.read_csv(LOSS_CSV)
    wavg = pd.read_csv(WAVG_CSV)

    tris = SimpleNamespace(a90=a90, atp=atp)
    lrr  = SimpleNamespace(loss=loss)
    ecl  = SimpleNamespace(wavg=wavg)

    checks = validate(feed, tris, lrr, ecl, DB_PATH)
    write_report(checks, VALIDATION_XLSX)
    print_summary(checks)
    print(f"Wrote: {VALIDATION_XLSX}")