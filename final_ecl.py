"""
FINAL ECL  -  disbursal-weighted average loss rate
==================================================
Per the mentor's spec, the final step is NOT a per-quarter ECL. It is a
WEIGHTED AVERAGE LOSS RATE over a chosen observation window:

    weighted_LR = SUMPRODUCT(loss_rate[q1:q2], disb_amt[q1:q2]) / SUM(disb_amt[q1:q2])

reproducing   =SUMPRODUCT(O125:O140,$B125:$B140)/SUM($B125:$B140)
where O = the per-quarter loss rate at the anchor and B = disbursal amount
(the weight, taken from the 90+ amount table).

Windows requested (only the FY range and the anchor change):
    FY20-Q1 .. FY23-Q4  @  84M
    FY16-Q1 .. FY23-Q4  @  84M
    FY16-Q1 .. FY23-Q4  @ 120M

Final step (mentor-confirmed): ECL = headline weighted_LR x window disbursal,
i.e. the total disbursal amount of the quarters INSIDE the headline window.
CURRENT_TPOS (latest observed outstanding, the triangle diagonal) is still
computed and shown for information only - it is NOT the exposure.

-------------------------------------------------------------------------------
This module exposes a PURE function:

    run(loss, atp) -> FinalECL(by_quarter, wavg, portfolio_tpos, portfolio_ecl)

`loss` is the per-quarter loss table (loss_rate.run().loss); `atp` is the TPOS
amount triangle (chain_ladder.run().atp). run() reads nothing and writes nothing.
CSV/Excel side effects live only in the `if __name__ == "__main__"` block below.
"""

import calendar
from datetime import date
from typing import NamedTuple

import numpy as np
import pandas as pd

from config import *      # AS_OF, MOB_LIST, WINDOWS, HEADLINE, fy_key


class FinalECL(NamedTuple):
    by_quarter: pd.DataFrame  # loss table + CURRENT_MOB + CURRENT_TPOS (was ecl_by_quarter.csv)
    wavg: pd.DataFrame        # one row per observation window (was weighted_loss_rate.csv)
    portfolio_tpos: float     # sum of current-TPOS diagonal (crores) - INFO ONLY, not the exposure
    ecl_pct: float            # THE ECL: headline window's weighted-avg loss rate (%)
    ecl_amount: float

# current exposure = diagonal of TPOS triangle
def quarter_end(label):
    fy = 2000 + int(label[2:4]); q = int(label[-1])
    y, m = {1: (fy - 1, 6), 2: (fy - 1, 9), 3: (fy - 1, 12), 4: (fy, 3)}[q]
    return date(y, m, calendar.monthrange(y, m)[1])


def current_mob(label, as_of=AS_OF):
    qe = quarter_end(label)
    months = (as_of.year - qe.year) * 12 + (as_of.month - qe.month)
    valid = [m for m in MOB_LIST if m <= months]
    return max(valid) if valid else -1


def run(loss: pd.DataFrame, atp: pd.DataFrame) -> FinalECL:
    """Compute per-window weighted-average loss rates and the provisional ECL. No I/O."""
    loss = loss.copy()
    atp = atp.copy(); atp.columns = [int(c) for c in atp.columns]

    loss["CURRENT_MOB"]  = loss.FY_QUARTER.map(current_mob)
    loss["CURRENT_TPOS"] = [
        atp.loc[q, cm] if (cm := current_mob(q)) in atp.columns else 0.0
        for q in loss.FY_QUARTER
    ]
    portfolio_tpos = loss.CURRENT_TPOS.sum()

    # WEIGHTED AVERAGE LOSS RATE PER WINDOW
    rows = []
    for label, q1, q2, A in WINDOWS:
        k1, k2 = fy_key(q1), fy_key(q2)
        win = loss[loss.FY_QUARTER.map(lambda q: k1 <= fy_key(q) <= k2)]
        lr, w = win[f"LOSS_RATE_{A}M"].to_numpy(), win.DISBURSAL_AMT.to_numpy()
        wavg = float(np.dot(lr, w) / w.sum()) if w.sum() else 0.0
        rows.append({"WINDOW": label, "FY_START": q1, "FY_END": q2, "ANCHOR_MOB": A,
                     "N_QUARTERS": len(win), "TOTAL_DISB": w.sum(),
                     "SIMPLE_AVG_LR": float(lr.mean()), "WEIGHTED_AVG_LR": wavg,
                     "ECL_PCT": wavg,                     # the ECL for this window
                     "ECL_AMT_DERIVED": wavg * w.sum()})  # info only
    wavg_df = pd.DataFrame(rows)

    headline = wavg_df[wavg_df.WINDOW == HEADLINE].iloc[0]
    ecl_pct    = float(headline.WEIGHTED_AVG_LR)          # reported figure
    ecl_amount = ecl_pct * float(headline.TOTAL_DISB)     # derived, info only

    return FinalECL(by_quarter=loss, wavg=wavg_df, portfolio_tpos=portfolio_tpos,
                    ecl_pct=ecl_pct, ecl_amount=ecl_amount)


# Standalone entrypoint: disk I/O + presentation-only Excel.
# None of this runs on import. `python final_ecl.py` writes the same outputs.
# The Excel writer is a candidate to move into report.py in a later step.
def _to_excel(res: FinalECL, path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wavg_df, loss = res.wavg, res.by_quarter
    portfolio_tpos, portfolio_ecl = res.portfolio_tpos, res.portfolio_ecl

    HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF")
    IFL = PatternFill("solid", fgColor="DDEBF7"); TOT = PatternFill("solid", fgColor="C6E0B4")
    WARN = PatternFill("solid", fgColor="FFC7CE")
    C = Alignment("center", "center"); BD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    CR, PC = "#,##0.0000", "0.00%"

    wb = Workbook(); ws = wb.active; ws.title = "Weighted_LossRate"
    heads = ["WINDOW", "FY_START", "FY_END", "ANCHOR", "N_QTRS", "TOTAL_DISB", "SIMPLE_AVG", "WEIGHTED_AVG"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(1, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
    for i, r in wavg_df.iterrows():
        rr = 2 + i
        ws.cell(rr, 1, r.WINDOW); ws.cell(rr, 2, r.FY_START); ws.cell(rr, 3, r.FY_END)
        ws.cell(rr, 4, int(r.ANCHOR_MOB)); ws.cell(rr, 5, int(r.N_QUARTERS))
        ws.cell(rr, 6, round(r.TOTAL_DISB, 4)).number_format = CR
        ws.cell(rr, 7, round(r.SIMPLE_AVG_LR, 6)).number_format = PC
        wc = ws.cell(rr, 8, round(r.WEIGHTED_AVG_LR, 6)); wc.number_format = PC; wc.font = Font(bold=True)
        if r.WEIGHTED_AVG_LR > 1: wc.fill = WARN            # implausible -> flag
        for cc in range(1, 9): ws.cell(rr, cc).alignment = C; ws.cell(rr, cc).border = BD
    r0 = 3 + len(wavg_df)
    ws.cell(r0, 1, "Headline window").font = Font(bold=True); ws.cell(r0, 2, HEADLINE)
    ws.cell(r0 + 1, 1, "Portfolio current TPOS (cr)").font = Font(bold=True)
    ws.cell(r0 + 1, 2, round(portfolio_tpos, 4)).number_format = CR
    ws.cell(r0 + 2, 1, "Portfolio ECL (cr)  [provisional]").font = Font(bold=True)
    ec = ws.cell(r0 + 2, 2, round(portfolio_ecl, 4)); ec.number_format = CR; ec.fill = TOT
    for col, w in zip("ABCDEFGH", [20, 10, 10, 9, 9, 14, 13, 14]):
        ws.column_dimensions[col].width = w

    # per-quarter detail
    ws2 = wb.create_sheet("Per_Quarter")
    heads2 = ["FY_QUARTER", "DISB_AMT", "LOSS_RATE_84M", "LOSS_RATE_120M", "CURRENT_MOB", "CURRENT_TPOS"]
    for j, h in enumerate(heads2, 1):
        c = ws2.cell(1, j, h); c.fill, c.font, c.alignment, c.border = HF, HFONT, C, BD
    for i, r in loss.iterrows():
        rr = 2 + i
        ws2.cell(rr, 1, r.FY_QUARTER).fill = IFL
        ws2.cell(rr, 2, round(r.DISBURSAL_AMT, 4)).number_format = CR
        c84 = ws2.cell(rr, 3, round(r.LOSS_RATE_84M, 6)); c84.number_format = PC
        c120 = ws2.cell(rr, 4, round(r.LOSS_RATE_120M, 6)); c120.number_format = PC
        if r.LOSS_RATE_120M > 1: c120.fill = WARN
        ws2.cell(rr, 5, int(r.CURRENT_MOB))
        ws2.cell(rr, 6, round(r.CURRENT_TPOS, 6)).number_format = CR
        for cc in range(1, 7): ws2.cell(rr, cc).alignment = C; ws2.cell(rr, cc).border = BD
    ws2.column_dimensions["A"].width = 12
    for col in "BCDEF": ws2.column_dimensions[col].width = 15
    wb.save(path)


def _print_summary(res: FinalECL) -> None:
    wavg_df, loss = res.wavg, res.by_quarter
    portfolio_tpos, portfolio_ecl = res.portfolio_tpos, res.portfolio_ecl
    headline = wavg_df[wavg_df.WINDOW == HEADLINE].iloc[0]

    print("=" * 68); print("FINAL ECL  -  weighted-average loss rate"); print("=" * 68)
    print(f"portfolio current TPOS : {portfolio_tpos:,.2f} cr")
    print(f"headline window disb  : {headline.TOTAL_DISB:,.2f} cr\n")
    print(f"{'WINDOW':<20}{'QTRS':>6}{'SIMPLE':>12}{'WEIGHTED':>14}   flag")
    for _, r in wavg_df.iterrows():
        flag = "  <-- IMPLAUSIBLE (>100%)" if r.WEIGHTED_AVG_LR > 1 else ""
        print(f"{r.WINDOW:<20}{int(r.N_QUARTERS):>6}{r.SIMPLE_AVG_LR:>11.4%}{r.WEIGHTED_AVG_LR:>14.4%}{flag}")

    # hand-check the headline window against an explicit SUMPRODUCT
    lab, q1, q2, A = [w for w in WINDOWS if w[0] == HEADLINE][0]
    k1, k2 = fy_key(q1), fy_key(q2)
    win = loss[loss.FY_QUARTER.map(lambda q: k1 <= fy_key(q) <= k2)]
    sp = sum(win[f"LOSS_RATE_{A}M"] * win.DISBURSAL_AMT); sw = win.DISBURSAL_AMT.sum()
    print(f"\nhand-check {HEADLINE} ({len(win)} qtrs, {q1}..{q2}):")
    print(f"    SUMPRODUCT = {sp:.6f}   SUM(disb) = {sw:.6f}   -> {sp/sw:.6%}")
    print(f"    engine     = {headline.WEIGHTED_AVG_LR:.6%}   match = {np.isclose(sp/sw, headline.WEIGHTED_AVG_LR)}")
    print(f"\nECL ({HEADLINE}) = weighted_LR x window disbursal = {portfolio_ecl:,.2f} cr"
          f"   (= {headline.WEIGHTED_AVG_LR:.4%} of {headline.TOTAL_DISB:,.2f} cr disbursal)")
    print(f"    [info only] portfolio current TPOS = {portfolio_tpos:,.2f} cr")


if __name__ == "__main__":
    loss = pd.read_csv("loss_rate.csv")
    atp  = pd.read_csv("tri_tpos_amt.csv", index_col=0); atp.columns = [int(c) for c in atp.columns]

    res = run(loss, atp)

    res.by_quarter.to_csv("ecl_by_quarter.csv", index=False)
    res.wavg.to_csv("weighted_loss_rate.csv", index=False)
    _to_excel(res, "ecl_summary.xlsx")

    _print_summary(res)
    print("\nWrote: weighted_loss_rate.csv, ecl_by_quarter.csv, ecl_summary.xlsx")